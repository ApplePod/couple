#!/usr/bin/env python3
"""영재·시온 월정산 엑셀 — 7월 1일 기준"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT = "/Users/harry/Desktop/사업문서/영재시온월정산/영재시온_월정산.xlsx"

MONEY = '#,##0"원"'
BOLD = Font(bold=True, size=11)
TITLE = Font(bold=True, size=16)
SECTION = Font(bold=True, size=12, color="333333")
NOTE = Font(size=10, color="888888")
THIN = Side(style="thin", color="DDDDDD")
BD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
YELLOW = PatternFill("solid", fgColor="FFF9E6")
GREEN = PatternFill("solid", fgColor="E8F5E9")
RED = PatternFill("solid", fgColor="FEF2F2")
BLUE_HDR = PatternFill("solid", fgColor="EEF3FF")
PURPLE_HDR = PatternFill("solid", fgColor="F3EEFF")

YJ_INCOME = 5124120
SN_INCOME = 5418470

YJ_ITEMS = [
    ("IRP", "ETF 장투 · 자동이체", 500000),
    ("연금저축펀드", "고영 ETF · 자동이체", 1000000),
    ("청년미래적금", "자동이체", 500000),
    ("국장 투자", "", 500000),
    ("주택청약", "자동이체", 250000),
    ("주거비", "", 750000),
    ("월세 대출이자", "", 130000),
    ("용돈", "", 700000),
    ("전기세", "예상", 40000),
    ("가스료", "예상", 30000),
    ("수도요금", "예상", 20000),
    ("인터넷", "예상", 35000),
    ("윌리엄 할부", "780만 무이자 12개월", 650000),
]

SN_ITEMS = [
    ("용돈", "", 700000),
    ("여행비", "국내2·해외1", 200000),
    ("ISA", "자동이체", 500000),
    ("IRP", "자동이체", 500000),
    ("연금저축펀드", "자동이체", 1000000),
    ("주택청약", "자동이체", 250000),
    ("청년도약적금", "자동이체", 700000),
    ("국장 투자", "", 300000),
    ("미국장 투자", "", 300000),
    ("코인 + 금", "", 200000),
]

ISA_ITEMS = [
    ("7월 목돈", 20000000),
]


def c(ws, r, col, val=None, formula=None, bold=False, fill=None, fmt=None, align="left", font=None):
    cl = ws.cell(row=r, column=col)
    cl.value = formula if formula else val
    cl.font = font or (BOLD if bold else Font(size=11))
    cl.border = BD
    cl.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if fmt:
        cl.number_format = fmt
    if fill:
        cl.fill = fill
    return cl


def write_person_block(ws, start_row, name, income, items, col_item, col_note, col_amt, hdr_fill):
    A = get_column_letter(col_amt)

    r = start_row
    c(ws, r, col_item, f"【 {name} 】  월 실수령", bold=True, fill=hdr_fill)
    c(ws, r, col_amt, income, fmt=MONEY, align="right", bold=True, fill=hdr_fill)
    income_row = r

    r += 1
    c(ws, r, col_item, "항목", bold=True, align="center")
    c(ws, r, col_note, "비고", bold=True, align="center")
    c(ws, r, col_amt, "월 금액", bold=True, align="center")

    data_start = r + 1
    for i, (item, note, amt) in enumerate(items):
        row = data_start + i
        c(ws, row, col_item, item)
        c(ws, row, col_note, note, font=NOTE)
        c(ws, row, col_amt, amt, fmt=MONEY, align="right")
    data_end = data_start + len(items) - 1

    sub_row = data_end + 1
    c(ws, sub_row, col_item, "소계", bold=True)
    c(ws, sub_row, col_amt, formula=f"=SUM({A}{data_start}:{A}{data_end})",
      fmt=MONEY, align="right", bold=True, fill=YELLOW)

    bal_row = sub_row + 1
    c(ws, bal_row, col_item, "남는 돈", bold=True, fill=GREEN)
    c(ws, bal_row, col_amt, formula=f"={A}{income_row}-{A}{sub_row}",
      fmt=MONEY, align="right", bold=True, fill=GREEN)

    return income_row, sub_row, bal_row


def build(ws):
    ws.title = "월정산"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 3
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 14

    c(ws, 1, 2, "영재 · 시온 월정산", bold=True)
    ws.merge_cells("B1:H1")
    ws["B1"].font = TITLE
    ws.merge_cells("B2:H2")
    c(ws, 2, 2, "7월 1일 시작 — 계좌 개설 및 입금 예정", font=NOTE, align="center")

    yj_inc, yj_sub, yj_bal = write_person_block(
        ws, 4, "영재", YJ_INCOME, YJ_ITEMS, 2, 3, 4, BLUE_HDR)
    sn_inc, sn_sub, sn_bal = write_person_block(
        ws, 4, "시온", SN_INCOME, SN_ITEMS, 6, 7, 8, PURPLE_HDR)

    # 가구 합계
    r = max(yj_bal, sn_bal) + 3
    c(ws, r, 2, "【 가구 합계 】", bold=True, font=SECTION)
    ws.merge_cells(f"B{r}:C{r}")

    r += 1
    c(ws, r, 2, "실수령 합계")
    c(ws, r, 4, formula=f"=D{yj_inc}+H{sn_inc}", fmt=MONEY, align="right", bold=True)

    r += 1
    c(ws, r, 2, "지출·저축 합계")
    c(ws, r, 4, formula=f"=D{yj_sub}+H{sn_sub}", fmt=MONEY, align="right", bold=True, fill=YELLOW)

    r += 1
    c(ws, r, 2, "남는 돈 합계", bold=True, fill=GREEN)
    c(ws, r, 4, formula=f"=D{yj_bal}+H{sn_bal}", fmt=MONEY, align="right", bold=True, fill=GREEN)

    # ISA
    r += 3
    c(ws, r, 2, "영재 ISA — 목돈 7월 일시불", bold=True, font=SECTION)
    ws.merge_cells(f"B{r}:D{r}")

    isa_start = r + 1
    c(ws, isa_start, 2, "일정", bold=True, align="center")
    c(ws, isa_start, 3, "내용", bold=True, align="center")
    c(ws, isa_start, 4, "금액", bold=True, align="center")

    for i, (sched, amt) in enumerate(ISA_ITEMS):
        row = isa_start + 1 + i
        c(ws, row, 2, sched)
        c(ws, row, 3, "목돈 2,000만원")
        c(ws, row, 4, amt, fmt=MONEY, align="right")

    isa_end = isa_start + len(ISA_ITEMS)
    c(ws, isa_end + 1, 2, "ISA 합계", bold=True)
    c(ws, isa_end + 1, 4, formula=f"=SUM(D{isa_start+1}:D{isa_end})",
      fmt=MONEY, align="right", bold=True, fill=YELLOW)

    # 안내
    r = isa_end + 3
    ws.merge_cells(f"B{r}:H{r+1}")
    ws[f"B{r}"] = (
        "※ 공과금(전기·가스·수도·인터넷)은 2인 투룸 12평 예상치\n"
        "※ 윌리엄 780만원 무이자 12개월 → 월 65만원 (영재)\n"
        "※ 영재 ISA 목돈 2천만원 — 7월 일시불 (월정산 합계 제외)"
    )
    ws[f"B{r}"].font = NOTE
    ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "B4"


def main():
    wb = Workbook()
    build(wb.active)
    wb.save(OUTPUT)
    print(f"생성 완료: {OUTPUT}")


if __name__ == "__main__":
    main()
